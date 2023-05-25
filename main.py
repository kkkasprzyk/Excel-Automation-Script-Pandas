import time
import subprocess
import numpy
import numpy as np
import pandas as pd
import self as self
from openpyxl import load_workbook
import psutil
import signal
import tkinter as tk
import os
from pprint import pprint
from openpyxl import load_workbook


def count_cells_in_column(file_path, column):
    wb = load_workbook(file_path)
    sheet = wb.active

    cell_count = 0
    row = 1

    while True:
        cell_value = sheet.cell(row=row, column=column).value
        if cell_value is None:
            break
        cell_count += 1
        row += 1

    return cell_count


# forced shutdown excel to complete saving
# dopiska:as/stability to filter , z PI jest DC tylko , POC to net bateryjny lub konektora , dekapling nie ma tam sensu , bat, kl15,kl30 ,poc  /// notatki co i jak !
###
# ############# os.system("taskkill /f /im  EXCEL.exe")  ##########################################
###
os.system("taskkill /f /im  EXCEL.exe")

# path to Analysis_Plan_Template.xlsx file
test_file = os.path.realpath('Iveco_Cluster_Ticket_List_xlsx.xlsx')
file_path_2 = os.path.realpath('Analysis_Plan_Iveco_Cluster.xlsx')
file_path = os.path.realpath('testowy_plik.xlsx')  ## plik testowy excela do uzupelnienia

# loading excel file to script and reading properly sheet
issues = pd.read_excel(test_file,sheet_name='Issues')
setup = pd.read_excel(file_path,sheet_name='Setup')
block_interface = pd.read_excel(file_path,sheet_name='Block | Interface')
block_type = pd.read_excel(file_path_2,sheet_name='Blocks')


# Opening excel file , path to  executable file excel and xlsx file
prog = r"c:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE" # path to exe file of excel , flexible path
# OpenIt = subprocess.Popen([prog, file])


# list_2 ma miec wielkosc ilosci analiz w excelu analysis plan template czyli ilosci wierszy
list_2 = np.empty([2,block_type.shape[0]],dtype=object)
optional = np.empty([2,block_type.shape[0]],dtype=object)
priority = np.empty([2,block_type.shape[0]],dtype=object)

# wb = load_workbook(file_path,data_only=True)
# sheet_setup = wb['Setup']
wb_issues = load_workbook(test_file)
sheet_issues = wb_issues['Issues']

# OpenIt.terminate()


for l in range(0, block_type.shape[0]):
    list_2[0][l] = str(block_type[l:l + 1]["Block name"].values)  ## list of block type
    print(list_2[0][l].replace('[', '').replace(']', '+'))
    for s in range(1,2):
        list_2[s][l] = block_type[l:l + 1]["Default analysis list"].str.split(', ', expand=True).values.tolist()
        optional[s][l] = block_type[l:l + 1]["Optional analysis list"].str.split(', ', expand=True).values.tolist()
        priority[s][l] = block_type[l:l + 1]["Priority"].str.split(', ', expand=True).values.tolist()
        # print(priority[s][l])
        # wyswietlanie ilosci analiz na jeden block name(POC,3v3)
        #   print(np.shape(list_2[1][l]),'+')
        #   print(list_2[1][0],'+')

count = 0
first_element=0
for k in range(0, block_type.shape[0]):
        print("Długość analiz_plan_template -->",np.shape(list_2[1][k])[1]) # wyświetlanie długości listy analiz z Analysis_Plan_Template
        wielkosc = np.shape(list_2[1][k])[1]
        for q in range(0,np.shape(list_2[1][k])[1]):
            print("q->",q)
            if list_2[1][k][0][q] == 'nic':
                # count=+2
                # sheet_issues.cell(row=1, column=7).value = "PI DC" + list_2[0][k].strip("[]'")
                print(list_2[1][k][0][q])
            if q == ((np.shape(list_2[1][k])[1])-1):
                # print(((np.shape(list_2[1][k])[1])),"tutaj q")
                # print(list_2[1][k][0],  "sprawdzamy gdzie to")
                count += wielkosc * 2 + 1
                print("Suma analiz - > ", count)
                # print("wielkosc - > ", wielkosc)
                # print("count - > ", count)
                # print("koniec")
                wb_issues.save(test_file)



list_1 = [
    [['WCCA', 'AC/Stability', 'PI', 'EMC (RE;RI;ESD)']],
    [['WCCA', 'PI,x']],
    [['WCCA', 'PI']],
    [['WCCA', 'PI']],
    [['WCCA']],
    [['SI']],
    [['S-params/TDR']],
    [['WCCA']],
    [['WCCA', 'PI']]
]


for sublist_index, sublist in enumerate(list_2):
    for analysis_list_index, analysis_list in enumerate(sublist):
        for analysis_index, analysis in enumerate(analysis_list):
            print(f'Iterating sublist index: {sublist_index}, analysis list index: {analysis_list_index}, analysis index: {analysis_index}, analysis: {analysis}')









my_list = [
  ["['POC']", "['3V3']", "['1V8']", "['1V0']", "['Serializer']", "['LVDS']", "['GMSL']", "['I2C']", "['Backlight converter']"],
  [
    list([['WCCA', 'AC/Stability', 'PI', 'EMC (RE;RI;ESD)']]),
    list([['WCCA', 'PI']]),
    list([['WCCA', 'PI']]),
    list([['WCCA', 'PI']]),
    list([['WCCA']]),
    list([['SI']]),
    list([['S-params/TDR']]),
    list([['WCCA']]),
    list([['WCCA', 'PI']])
  ]
]

for i in range(len(list_2[0])):
    poc = my_list[0][i].strip("[]'")
    analyses = my_list[1][i]

    print(poc)
    for analysis in analyses:
        for element in analysis:
            print(element)
    print()  # Dodaj pustą linię między elementami
# open excel


row_index = 2

for i in range(len(my_list[0])):
    poc = my_list[0][i].strip("[]'")
    analyses = my_list[1][i]
    sheet_issues.cell(row=row_index, column=2).value = "NAZWA PROJEKTU"
    sheet_issues.cell(row=row_index, column=7).value = poc
    row_index += 1
    print(len(analyses[0]))
    print("pauza")
    for analysis in analyses:
        print("coś")
        for element in analysis:
            sheet_issues.cell(row=row_index, column=2).value = "NAZWA PROJEKTU"
            if element == 'PI':
                actions = ["AC", "AC REVIEW", "DC", "DC REVIEW"]
                for action in actions:
                    sheet_issues.cell(row=row_index, column=2).value = "NAZWA PROJEKTU"
                    sheet_issues.cell(row=row_index, column=7).value = element + " " + my_list[0][i].strip("[]'") + " " + action
                    row_index += 1
            else:
                sheet_issues.cell(row=row_index, column=2).value = "NAZWA PROJEKTU"
                sheet_issues.cell(row=row_index, column=4).value = ""
                sheet_issues.cell(row=row_index, column=7).value = element + " " + my_list[0][i].strip("[]'")
                row_index += 1
                sheet_issues.cell(row=row_index, column=2).value = "NAZWA PROJEKTU"
                sheet_issues.cell(row=row_index, column=7).value = element + " " + my_list[0][i].strip("[]'") + " REVIEW"
                row_index += 1



wb_issues.save(test_file)
subprocess.Popen([prog, test_file])
pprint(list_2)



# save excel


# print(list_2[1][0:10][0:10])
# if count == 0:
#     sheet_issues.cell(row=2, column=7).value = list_2[0][k].strip("[]'")
#     wb_issues.save(test_file)
# elif count > 0:
#     sheet_issues.cell(row=count + 2, column=7).value = list_2[0][k].strip("[]'")